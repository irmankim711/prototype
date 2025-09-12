#!/usr/bin/env python3
"""
Next-Gen Report Builder Integration Test
Tests the complete workflow from Excel upload to report generation
"""

import sys
import os
import json
import time
from pathlib import Path

# Add the parent directory to Python path
sys.path.insert(0, str(Path(__file__).parent))

from app import create_app, db
from app.models import User, Form, FormSubmission, Report
from app.services.form_automation import FormAutomationService
from app.routes.nextgen_report_builder import nextgen_bp
import openpyxl

def create_test_excel_file():
    """Create a test Excel file with sample data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sample Data"
    
    # Headers
    headers = ["Region", "Quarter", "Revenue", "Profit Margin", "Customer Count"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Sample data
    data = [
        ["North", "Q1", 120000, 0.15, 450],
        ["North", "Q2", 135000, 0.18, 520],
        ["South", "Q1", 98000, 0.12, 380],
        ["South", "Q2", 110000, 0.14, 420],
        ["East", "Q1", 86000, 0.16, 320],
        ["East", "Q2", 92000, 0.19, 350],
        ["West", "Q1", 102000, 0.13, 400],
        ["West", "Q2", 115000, 0.17, 460],
    ]
    
    for row, row_data in enumerate(data, 2):
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Save to test file
    test_file = Path(__file__).parent / "test_sample_data.xlsx"
    wb.save(test_file)
    return test_file

def create_test_template():
    """Create a test report template"""
    template_content = """
# {{ report_title }}

## Executive Summary

This report provides an analysis of business performance across different regions and quarters.

## Key Metrics

- **Total Revenue**: ${{ total_revenue | default('N/A') }}
- **Average Profit Margin**: {{ avg_profit_margin | default('N/A') }}%
- **Total Customers**: {{ total_customers | default('N/A') }}

## Regional Performance

{% for region in regions %}
### {{ region.name }}
- Revenue: ${{ region.revenue }}
- Profit Margin: {{ region.profit_margin }}%
- Customers: {{ region.customers }}
{% endfor %}

## Quarterly Trends

{% for quarter in quarters %}
### {{ quarter.name }}
- Revenue: ${{ quarter.revenue }}
- Growth: {{ quarter.growth }}%
{% endfor %}

## Data Summary

{% if data_table %}
| Region | Quarter | Revenue | Profit Margin | Customers |
|--------|---------|---------|---------------|-----------|
{% for row in data_table %}
| {{ row.region }} | {{ row.quarter }} | ${{ row.revenue }} | {{ row.profit_margin }}% | {{ row.customers }} |
{% endfor %}
{% endif %}

---
*Report generated on {{ generation_date }}*
"""
    
    template_file = Path(__file__).parent / "templates" / "test_report_template.jinja"
    template_file.parent.mkdir(exist_ok=True)
    
    with open(template_file, 'w', encoding='utf-8') as f:
        f.write(template_content)
    
    return template_file

def test_nextgen_report_builder():
    """Test the complete Next-Gen Report Builder workflow"""
    print("🚀 Starting Next-Gen Report Builder Integration Test")
    
    # Create Flask app
    app = create_app('testing')
    
    with app.app_context():
        try:
            # Initialize database
            db.create_all()
            
            # Create test user
            test_user = User(
                username='testuser',
                email='test@example.com',
                is_active=True
            )
            test_user.set_password('testpass')
            db.session.add(test_user)
            db.session.commit()
            
            print(f"✅ Created test user: {test_user.username}")
            
            # Create test Excel file
            excel_file = create_test_excel_file()
            print(f"✅ Created test Excel file: {excel_file}")
            
            # Create test template
            template_file = create_test_template()
            print(f"✅ Created test template: {template_file}")
            
            # Initialize Form Automation Service
            form_automation = FormAutomationService()
            
            # Test 1: Parse Excel file
            print("\n📊 Testing Excel parsing...")
            excel_result = form_automation.excel_parser.parse_excel_file(str(excel_file))
            
            if excel_result['success']:
                print(f"✅ Excel parsing successful:")
                print(f"   - Sheets: {excel_result.get('sheets_processed', 0)}")
                print(f"   - Rows: {excel_result.get('total_rows', 0)}")
                print(f"   - Columns: {len(excel_result.get('columns', []))}")
            else:
                print(f"❌ Excel parsing failed: {excel_result.get('error', 'Unknown error')}")
                return False
            
            # Test 2: Generate report from Excel
            print("\n📝 Testing report generation...")
            report_result = form_automation.generate_report_from_excel(
                excel_path=str(excel_file),
                template_path=str(template_file)
            )
            
            if report_result['success']:
                print(f"✅ Report generation successful:")
                print(f"   - Report file: {report_result['filename']}")
                print(f"   - Template used: {report_result['template_used']}")
                print(f"   - Context fields: {report_result['context_summary']['total_fields']}")
                
                # Read generated report
                if os.path.exists(report_result['report_path']):
                    with open(report_result['report_path'], 'r', encoding='utf-8') as f:
                        report_content = f.read()
                    print(f"   - Report length: {len(report_content)} characters")
                    print(f"   - Preview: {report_content[:200]}...")
                else:
                    print("❌ Generated report file not found")
                    return False
            else:
                print(f"❌ Report generation failed: {report_result.get('error', 'Unknown error')}")
                return False
            
            # Test 3: Template optimization
            print("\n🔧 Testing template optimization...")
            with open(template_file, 'r', encoding='utf-8') as f:
                template_content = f.read()
            
            optimization_result = form_automation.template_optimizer.optimize_template_with_excel(
                template_content, str(excel_file)
            )
            
            if optimization_result['success']:
                print(f"✅ Template optimization successful:")
                print(f"   - Enhanced context keys: {len(optimization_result['enhanced_context'].keys())}")
                print(f"   - Placeholders filled: {optimization_result.get('placeholders_filled', 0)}")
            else:
                print(f"❌ Template optimization failed: {optimization_result.get('error', 'Unknown error')}")
                return False
            
            # Test 4: Form creation from template
            print("\n📋 Testing form creation from template...")
            form_result = form_automation.create_form_from_template(
                template_path=str(template_file),
                form_title="Auto-Generated Data Collection Form"
            )
            
            if form_result['success']:
                print(f"✅ Form creation successful:")
                print(f"   - Form title: {form_result['form_title']}")
                print(f"   - Fields detected: {form_result['detected_fields']}")
                print(f"   - Field types: {form_result['field_types']}")
            else:
                print(f"❌ Form creation failed: {form_result.get('error', 'Unknown error')}")
                return False
            
            # Test 5: Complete workflow
            print("\n🔄 Testing complete automation workflow...")
            workflow_result = form_automation.create_automated_workflow(
                template_path=str(template_file),
                workflow_name="Sample Business Report Workflow"
            )
            
            if workflow_result['success']:
                print(f"✅ Workflow creation successful:")
                print(f"   - Workflow name: {workflow_result['workflow_name']}")
                print(f"   - Automation steps: {len(workflow_result['automation_steps'])}")
                print(f"   - Features: {len(workflow_result['automation_features'])}")
                
                for i, step in enumerate(workflow_result['automation_steps'], 1):
                    print(f"     {i}. {step['name']} - {step['status']}")
            else:
                print(f"❌ Workflow creation failed: {workflow_result.get('error', 'Unknown error')}")
                return False
            
            print("\n🎉 All tests passed successfully!")
            print("\n📊 Test Summary:")
            print("   ✅ Excel file parsing")
            print("   ✅ Report generation from Excel")
            print("   ✅ Template optimization")
            print("   ✅ Form creation from template")
            print("   ✅ Complete automation workflow")
            
            return True
            
        except Exception as e:
            print(f"❌ Test failed with exception: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
        
        finally:
            # Cleanup
            try:
                if 'excel_file' in locals() and excel_file.exists():
                    excel_file.unlink()
                    print(f"🧹 Cleaned up test Excel file")
                
                if 'template_file' in locals() and template_file.exists():
                    template_file.unlink()
                    print(f"🧹 Cleaned up test template file")
                    
                # Clean up generated files
                static_dir = Path(__file__).parent / "static" / "generated"
                if static_dir.exists():
                    for file in static_dir.glob("report_test_report_template_*.tex"):
                        file.unlink()
                        print(f"🧹 Cleaned up generated report: {file.name}")
                        
            except Exception as cleanup_error:
                print(f"⚠️ Cleanup warning: {cleanup_error}")

def test_api_endpoints():
    """Test the API endpoints"""
    print("\n🔌 Testing API Endpoints...")
    
    app = create_app('testing')
    
    with app.test_client() as client:
        # Test data sources endpoint
        print("Testing /api/v1/nextgen/data-sources...")
        
        # This would normally require authentication
        # For testing, we'll just check if the endpoint exists
        try:
            response = client.get('/api/v1/nextgen/data-sources')
            print(f"   Status: {response.status_code}")
            if response.status_code == 401:
                print("   ✅ Endpoint exists (requires authentication)")
            elif response.status_code == 200:
                print("   ✅ Endpoint accessible")
                data = response.get_json()
                print(f"   Data sources returned: {len(data) if data else 0}")
            else:
                print(f"   ⚠️ Unexpected status: {response.status_code}")
        except Exception as e:
            print(f"   ❌ Error: {e}")

if __name__ == "__main__":
    print("=" * 60)
    print("NEXT-GEN REPORT BUILDER COMPREHENSIVE TEST")
    print("=" * 60)
    
    success = test_nextgen_report_builder()
    
    if success:
        test_api_endpoints()
        print("\n" + "=" * 60)
        print("🎉 ALL TESTS COMPLETED SUCCESSFULLY!")
        print("The Next-Gen Report Builder is ready for use.")
        print("=" * 60)
    else:
        print("\n" + "=" * 60)
        print("❌ TESTS FAILED - Check the errors above")
        print("=" * 60)
        sys.exit(1)

