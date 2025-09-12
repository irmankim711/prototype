#!/usr/bin/env python3
"""
Direct Excel File Processing Test
Tests if backend can read Excel files directly without Celery/Flask overhead
"""

import os
import sys
import traceback
sys.path.append('.')

def test_excel_libraries():
    """Test if Excel processing libraries are available"""
    print("📚 Testing Excel Processing Libraries...")
    
    # Test openpyxl
    try:
        import openpyxl
        print("✅ openpyxl available:", openpyxl.__version__)
    except ImportError as e:
        print("❌ openpyxl not available:", e)
        return False
    
    # Test pandas
    try:
        import pandas as pd
        print("✅ pandas available:", pd.__version__)
    except ImportError as e:
        print("❌ pandas not available:", e)
        print("⚠️  pandas optional but recommended")
    
    # Test xlsxwriter (for writing)
    try:
        import xlsxwriter
        print("✅ xlsxwriter available:", xlsxwriter.__version__)
    except ImportError as e:
        print("❌ xlsxwriter not available:", e)
        print("⚠️  xlsxwriter optional but recommended")
    
    return True

def test_file_path_access():
    """Test file path access and permissions"""
    print("\n📁 Testing File Path Access...")
    
    # Test various potential paths
    test_paths = [
        r"C:\Users\IRMAN\OneDrive\Desktop\prototype\backend\app\static\uploads\excel\8d58ebb4-f8c0-4d83-82cd-ed5cc06c3009_ea29dfb58d314325860a32079b958c3a_SENARAI SEMAK PUNCAK ALAM.xlsx",
        r"C:\Users\IRMAN\OneDrive\Desktop\prototype\backend\static\uploads",
        r"C:\Users\IRMAN\OneDrive\Desktop\prototype\backend\app\static",
        r"C:\Users\IRMAN\OneDrive\Desktop\prototype\backend"
    ]
    
    for path in test_paths:
        if os.path.exists(path):
            print(f"✅ Path exists: {path}")
            if os.path.isfile(path):
                size = os.path.getsize(path)
                print(f"   📄 File size: {size:,} bytes")
            elif os.path.isdir(path):
                try:
                    files = os.listdir(path)
                    print(f"   📂 Directory contains {len(files)} items")
                    if files:
                        print(f"   📋 Sample files: {files[:3]}")
                except PermissionError:
                    print(f"   🚫 Permission denied to list directory")
        else:
            print(f"❌ Path missing: {path}")
    
    return True

def test_direct_excel_read():
    """Test direct Excel file reading"""
    print("\n📊 Testing Direct Excel File Reading...")
    
    # Test file path - update this with actual path found in your system
    file_path = r"C:\Users\IRMAN\OneDrive\Desktop\prototype\backend\app\static\uploads\excel\8d58ebb4-f8c0-4d83-82cd-ed5cc06c3009_ea29dfb58d314325860a32079b958c3a_SENARAI SEMAK PUNCAK ALAM.xlsx"
    
    # Also check for any .xlsx files in uploads directory
    uploads_dir = r"C:\Users\IRMAN\OneDrive\Desktop\prototype\backend\app\static\uploads\excel"
    if os.path.exists(uploads_dir):
        print(f"📂 Checking uploads directory: {uploads_dir}")
        try:
            files = [f for f in os.listdir(uploads_dir) if f.endswith('.xlsx')]
            if files:
                print(f"📋 Found {len(files)} Excel files:")
                for file in files[:5]:  # Show first 5
                    print(f"   - {file}")
                # Use the first found file for testing
                file_path = os.path.join(uploads_dir, files[0])
                print(f"🎯 Using file for testing: {file_path}")
            else:
                print("❌ No Excel files found in uploads directory")
        except Exception as e:
            print(f"❌ Error reading uploads directory: {e}")
    
    if not os.path.exists(file_path):
        print(f"❌ Test file not found: {file_path}")
        return False
    
    print(f"✅ Test file exists: {file_path}")
    print(f"📏 File size: {os.path.getsize(file_path):,} bytes")
    
    # Test with openpyxl
    try:
        import openpyxl
        print("\n🔧 Testing with openpyxl...")
        
        wb = openpyxl.load_workbook(file_path, read_only=True)
        print("✅ Workbook loaded successfully")
        print(f"📋 Available sheets: {wb.sheetnames}")
        
        # Read first sheet
        if wb.sheetnames:
            sheet_name = wb.sheetnames[0]
            sheet = wb[sheet_name]
            print(f"📊 Active sheet: {sheet_name}")
            print(f"📐 Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
            
            # Read first few rows
            print("🔍 First 3 rows of data:")
            for row_num in range(1, min(4, sheet.max_row + 1)):
                row_data = []
                for col_num in range(1, min(6, sheet.max_column + 1)):  # First 5 columns
                    cell = sheet.cell(row=row_num, column=col_num)
                    row_data.append(str(cell.value)[:50] if cell.value else "")
                print(f"   Row {row_num}: {row_data}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ openpyxl failed: {e}")
        print("🔍 Full traceback:")
        traceback.print_exc()
        return False
    
    # Test with pandas if available
    try:
        import pandas as pd
        print("\n🔧 Testing with pandas...")
        
        df = pd.read_excel(file_path, sheet_name=0, nrows=5)  # Read first 5 rows
        print("✅ Pandas read successful")
        print(f"📊 DataFrame shape: {df.shape}")
        print(f"📋 Columns: {list(df.columns)}")
        print("🔍 First few rows:")
        print(df.head().to_string())
        
    except ImportError:
        print("ℹ️  Pandas not available, skipping pandas test")
    except Exception as e:
        print(f"❌ Pandas failed: {e}")
        print("🔍 Full traceback:")
        traceback.print_exc()
    
    return True

def test_file_creation_permissions():
    """Test if we can create files in the expected directories"""
    print("\n📝 Testing File Creation Permissions...")
    
    test_dirs = [
        r"C:\Users\IRMAN\OneDrive\Desktop\prototype\backend\app\static\uploads",
        r"C:\Users\IRMAN\OneDrive\Desktop\prototype\backend\static\generated",
        r"C:\Users\IRMAN\OneDrive\Desktop\prototype\backend\app\static\generated"
    ]
    
    for test_dir in test_dirs:
        try:
            os.makedirs(test_dir, exist_ok=True)
            test_file = os.path.join(test_dir, "test_write.txt")
            
            with open(test_file, 'w') as f:
                f.write("Test write permission")
            
            if os.path.exists(test_file):
                os.remove(test_file)
                print(f"✅ Write permission OK: {test_dir}")
            else:
                print(f"❌ Write test failed: {test_dir}")
                
        except Exception as e:
            print(f"❌ Cannot write to {test_dir}: {e}")

def main():
    """Run all Excel processing tests"""
    print("🚀 Direct Excel Processing Test Suite")
    print("=" * 60)
    
    tests = [
        ("Library Availability", test_excel_libraries),
        ("File Path Access", test_file_path_access),
        ("Direct Excel Reading", test_direct_excel_read),
        ("File Creation Permissions", test_file_creation_permissions)
    ]
    
    results = {}
    
    for test_name, test_func in tests:
        print(f"\n{'='*20} {test_name} {'='*20}")
        try:
            result = test_func()
            results[test_name] = result if result is not None else True
        except Exception as e:
            print(f"💥 Test crashed: {e}")
            traceback.print_exc()
            results[test_name] = False
    
    # Summary
    print("\n" + "=" * 60)
    print("📊 EXCEL PROCESSING TEST RESULTS:")
    print("=" * 60)
    
    for test_name, result in results.items():
        status = "✅ PASS" if result else "❌ FAIL"
        print(f"  {status}: {test_name}")
    
    passed = sum(1 for r in results.values() if r)
    total = len(results)
    
    print(f"\n🎯 Overall: {passed}/{total} tests passed")
    
    if passed == total:
        print("🎉 All Excel processing tests passed!")
        print("📝 Excel files can be read directly - issue is likely in Celery or route handling")
    else:
        print("⚠️  Some Excel processing tests failed")
        print("📝 Fix library/path issues before testing Celery/routes")
    
    return passed == total

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)