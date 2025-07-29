"""
Form-to-Excel-to-Report Automation Service

This service provides comprehensive automation for:
1. Dynamic form creation based on templates
2. Form data export to Excel/Google Sheets
3. Automatic report generation from collected data
"""

import os
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple
import logging
from ..models import db, Form, FormSubmission
from .template_optimizer import TemplateOptimizerService
from .excel_parser import ExcelParserService
from jinja2 import Template
import uuid

# Optional Google Sheets support
try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSHEETS_AVAILABLE = True
except ImportError:
    GSHEETS_AVAILABLE = False
    logging.warning("Google Sheets integration not available. Install gspread and google-auth packages for full functionality.")

logger = logging.getLogger(__name__)


class FormAutomationService:
    """
    Complete automation service for form-to-report workflow
    """
    
    def __init__(self):
        self.template_optimizer = TemplateOptimizerService()
        self.excel_parser = ExcelParserService()
        
    def create_form_from_template(self, template_path: str, form_title: str = None) -> Dict[str, Any]:
        """
        Create a dynamic form based on template placeholders
        
        Args:
            template_path: Path to the template file
            form_title: Title for the generated form
            
        Returns:
            Dict containing form schema and metadata
        """
        try:
            logger.info(f"Creating form from template: {template_path}")
            
            # Read template content
            with open(template_path, 'r', encoding='utf-8') as f:
                template_content = f.read()
            
            # Extract placeholders from template
            from .template_converter import TemplateConverter
            syntax_type = TemplateConverter.analyze_template_syntax(template_content)
            
            if syntax_type in ['mustache', 'mixed']:
                template_content = TemplateConverter.mustache_to_jinja2(template_content)
            
            # Analyze template to extract required fields
            placeholders = self._extract_form_fields_from_template(template_content)
            
            # Generate form schema
            form_schema = self._generate_form_schema_from_placeholders(placeholders)
            
            # Set form title
            if not form_title:
                form_title = f"Data Collection Form - {os.path.basename(template_path)}"
            
            result = {
                'success': True,
                'form_title': form_title,
                'form_schema': form_schema,
                'template_path': template_path,
                'detected_fields': len(placeholders),
                'field_types': self._categorize_fields(placeholders)
            }
            
            logger.info(f"Successfully created form schema with {len(placeholders)} fields")
            return result
            
        except Exception as e:
            logger.error(f"Error creating form from template: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _extract_form_fields_from_template(self, template_content: str) -> List[Dict[str, Any]]:
        """Extract and categorize form fields from template placeholders"""
        import re
        
        placeholders = []
        seen_fields = set()
        
        # Pattern for Jinja2 variables
        variable_pattern = r'\{\{\s*([^}]+)\s*\}\}'
        
        for match in re.finditer(variable_pattern, template_content):
            field_path = match.group(1).strip()
            
            # Skip loop variables and complex expressions
            if any(skip in field_path for skip in ['item.', 'loop.', 'range(', 'if ', 'else']):
                continue
            
            # Handle nested fields
            field_parts = field_path.split('.')
            field_name = field_parts[-1] if len(field_parts) > 1 else field_path
            field_category = field_parts[0] if len(field_parts) > 1 else 'general'
            
            # Avoid duplicates
            field_key = f"{field_category}.{field_name}"
            if field_key in seen_fields:
                continue
            seen_fields.add(field_key)
            
            # Determine field type based on name patterns
            field_type = self._infer_field_type(field_name)
            
            placeholder = {
                'name': field_name,
                'full_path': field_path,
                'category': field_category,
                'type': field_type,
                'required': True,  # Default to required
                'label': self._generate_field_label(field_name),
                'description': self._generate_field_description(field_name, field_category)
            }
            
            placeholders.append(placeholder)
        
        return placeholders
    
    def _infer_field_type(self, field_name: str) -> str:
        """Infer form field type based on field name patterns"""
        field_name_lower = field_name.lower()
        
        # Email fields
        if any(term in field_name_lower for term in ['email', 'mail']):
            return 'email'
        
        # Phone fields
        if any(term in field_name_lower for term in ['phone', 'tel', 'mobile']):
            return 'phone'
        
        # Date fields
        if any(term in field_name_lower for term in ['date', 'day', 'month', 'year', 'birthday']):
            return 'date'
        
        # Time fields
        if any(term in field_name_lower for term in ['time', 'hour', 'minute']):
            return 'time'
        
        # Number fields
        if any(term in field_name_lower for term in ['number', 'count', 'total', 'amount', 'price', 'cost', 'quantity']):
            return 'number'
        
        # Rating fields
        if any(term in field_name_lower for term in ['rating', 'score', 'rank']):
            return 'rating'
        
        # Long text fields
        if any(term in field_name_lower for term in ['description', 'comment', 'note', 'feedback', 'detail', 'background', 'objective']):
            return 'textarea'
        
        # URL fields
        if any(term in field_name_lower for term in ['url', 'website', 'link']):
            return 'url'
        
        # Location fields
        if any(term in field_name_lower for term in ['location', 'address', 'place', 'venue']):
            return 'location'
        
        # Default to text
        return 'text'
    
    def _generate_field_label(self, field_name: str) -> str:
        """Generate user-friendly label from field name"""
        # Replace underscores and camelCase
        label = field_name.replace('_', ' ')
        
        # Handle camelCase
        import re
        label = re.sub(r'([A-Z])', r' \\1', label)
        
        # Capitalize words
        label = ' '.join(word.capitalize() for word in label.split())
        
        return label.strip()
    
    def _generate_field_description(self, field_name: str, category: str) -> str:
        """Generate helpful description for form field"""
        descriptions = {
            'program': f"Enter the {field_name} for the program",
            'participant': f"Provide participant {field_name} information",
            'evaluation': f"Rate or evaluate {field_name}",
            'tentative': f"Schedule information for {field_name}",
            'signature': f"Signature details for {field_name}",
            'general': f"Please provide {field_name} information"
        }
        
        return descriptions.get(category, descriptions['general'])
    
    def _generate_form_schema_from_placeholders(self, placeholders: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Generate complete form schema from extracted placeholders"""
        
        # Group fields by category
        categories = {}
        for placeholder in placeholders:
            category = placeholder['category']
            if category not in categories:
                categories[category] = []
            categories[category].append(placeholder)
        
        fields = []
        field_order = 0
        
        # Process each category
        for category_name, category_fields in categories.items():
            # Add section header
            if len(categories) > 1:  # Only add headers if multiple categories
                fields.append({
                    'id': f"section_{category_name}",
                    'type': 'section_header',
                    'label': category_name.title() + " Information",
                    'description': f"Please fill in the {category_name} details below",
                    'order': field_order,
                    'required': False
                })
                field_order += 1
            
            # Add fields for this category
            for field in category_fields:
                form_field = {
                    'id': f"{category_name}_{field['name']}",
                    'label': field['label'],
                    'type': field['type'],
                    'required': field['required'],
                    'description': field['description'],
                    'order': field_order,
                    'data_path': field['full_path'],  # For mapping back to template
                    'category': category_name
                }
                
                # Add field-specific configurations
                if field['type'] == 'rating':
                    form_field['validation'] = {'min': 1, 'max': 5}
                elif field['type'] == 'textarea':
                    form_field['validation'] = {'maxLength': 1000}
                elif field['type'] == 'text':
                    form_field['validation'] = {'maxLength': 255}
                elif field['type'] == 'email':
                    form_field['validation'] = {'pattern': r'^[^\\s@]+@[^\\s@]+\\.[^\\s@]+$'}
                elif field['type'] == 'phone':
                    form_field['validation'] = {'pattern': r'^[\\+]?[1-9]?[0-9]{7,15}$'}
                
                fields.append(form_field)
                field_order += 1
        
        return {
            'fields': fields,
            'settings': {
                'theme': 'modern',
                'submitButtonText': 'Submit Data',
                'showProgressBar': True,
                'allowMultipleSubmissions': True,
                'autoSave': True
            }
        }
    
    def _categorize_fields(self, placeholders: List[Dict[str, Any]]) -> Dict[str, int]:
        """Categorize fields by type for summary"""
        categories = {}
        for field in placeholders:
            field_type = field['type']
            categories[field_type] = categories.get(field_type, 0) + 1
        return categories
    
    def export_form_data_to_excel(self, form_id: int, output_path: str = None, include_analytics: bool = True) -> Dict[str, Any]:
        """
        Export form submission data to Excel with formatting and analytics
        
        Args:
            form_id: ID of the form to export
            output_path: Path where to save the Excel file
            include_analytics: Whether to include analytics sheet
            
        Returns:
            Dict with export results and file path
        """
        try:
            logger.info(f"Exporting form data to Excel for form ID: {form_id}")
            
            # Get form and submissions
            form = Form.query.get(form_id)
            if not form:
                return {'success': False, 'error': 'Form not found'}
            
            submissions = FormSubmission.query.filter_by(form_id=form_id).all()
            
            if not submissions:
                return {'success': False, 'error': 'No submissions found for this form'}
            
            # Prepare data
            data_rows = []
            for submission in submissions:
                row = {
                    'Submission ID': submission.id,
                    'Submitted At': submission.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'IP Address': submission.ip_address or 'N/A'
                }
                
                # Add form field data
                for field_name, value in submission.data.items():
                    row[field_name] = value
                
                data_rows.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(data_rows)
            
            # Generate output path if not provided
            if not output_path:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"form_data_{form.id}_{timestamp}.xlsx"
                output_dir = os.path.join(os.path.dirname(__file__), '../../static/exports')
                os.makedirs(output_dir, exist_ok=True)
                output_path = os.path.join(output_dir, filename)
            
            # Create Excel file with formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Write main data
                df.to_excel(writer, sheet_name='Form Data', index=False)
                
                # Format the main sheet
                workbook = writer.book
                worksheet = workbook['Form Data']
                self._format_excel_sheet(worksheet, df)
                
                # Add analytics sheet if requested
                if include_analytics:
                    analytics_data = self._generate_form_analytics(form, submissions)
                    analytics_df = pd.DataFrame(analytics_data)
                    analytics_df.to_excel(writer, sheet_name='Analytics', index=False)
                    
                    # Format analytics sheet
                    analytics_sheet = workbook['Analytics']
                    self._format_excel_sheet(analytics_sheet, analytics_df)
                
                # Add form schema sheet
                schema_data = self._format_form_schema_for_excel(form.schema)
                schema_df = pd.DataFrame(schema_data)
                schema_df.to_excel(writer, sheet_name='Form Schema', index=False)
                
                schema_sheet = workbook['Form Schema']
                self._format_excel_sheet(schema_sheet, schema_df)
            
            result = {
                'success': True,
                'file_path': output_path,
                'filename': os.path.basename(output_path),
                'total_submissions': len(submissions),
                'export_timestamp': datetime.now().isoformat(),
                'form_title': form.title
            }
            
            logger.info(f"Successfully exported {len(submissions)} submissions to {output_path}")
            return result
            
        except Exception as e:
            logger.error(f"Error exporting form data to Excel: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _format_excel_sheet(self, worksheet, df):
        """Apply professional formatting to Excel sheet"""
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply header formatting
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    def _generate_form_analytics(self, form: Form, submissions: List[FormSubmission]) -> List[Dict[str, Any]]:
        """Generate analytics data for the form"""
        if not submissions:
            return []
        
        analytics = []
        
        # Basic statistics
        analytics.append({
            'Metric': 'Total Submissions',
            'Value': len(submissions),
            'Description': 'Total number of form submissions'
        })
        
        # Submission timeline
        submission_dates = [s.submitted_at.date() for s in submissions]
        unique_dates = len(set(submission_dates))
        
        analytics.append({
            'Metric': 'Submission Days',
            'Value': unique_dates,
            'Description': 'Number of unique days with submissions'
        })
        
        # Average per day
        if unique_dates > 0:
            avg_per_day = len(submissions) / unique_dates
            analytics.append({
                'Metric': 'Average per Day',
                'Value': round(avg_per_day, 2),
                'Description': 'Average submissions per day'
            })
        
        # Date range
        if submissions:
            first_submission = min(s.submitted_at for s in submissions)
            last_submission = max(s.submitted_at for s in submissions)
            
            analytics.append({
                'Metric': 'First Submission',
                'Value': first_submission.strftime('%Y-%m-%d %H:%M'),
                'Description': 'Date and time of first submission'
            })
            
            analytics.append({
                'Metric': 'Last Submission',
                'Value': last_submission.strftime('%Y-%m-%d %H:%M'),
                'Description': 'Date and time of last submission'
            })
        
        return analytics
    
    def _format_form_schema_for_excel(self, schema: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Format form schema for Excel export"""
        schema_data = []
        
        for field in schema.get('fields', []):
            schema_data.append({
                'Field ID': field.get('id', ''),
                'Label': field.get('label', ''),
                'Type': field.get('type', ''),
                'Required': field.get('required', False),
                'Description': field.get('description', ''),
                'Category': field.get('category', ''),
                'Data Path': field.get('data_path', '')
            })
        
        return schema_data
    
    def generate_report_from_excel(self, excel_path: str, template_path: str, output_path: str = None) -> Dict[str, Any]:
        """
        Generate report from Excel data using template
        
        Args:
            excel_path: Path to Excel file with form data
            template_path: Path to report template
            output_path: Path for generated report
            
        Returns:
            Dict with generation results
        """
        try:
            logger.info(f"Generating report from Excel: {excel_path} using template: {template_path}")
            
            # Extract data from Excel
            extracted_data = self.excel_parser.parse_excel_file(excel_path)
            
            if not extracted_data['success']:
                return {
                    'success': False,
                    'error': f"Failed to parse Excel file: {extracted_data['error']}"
                }
            
            # Optimize template with Excel data
            with open(template_path, 'r', encoding='utf-8') as f:
                template_content = f.read()
            
            optimization_result = self.template_optimizer.optimize_template_with_excel(
                template_content, excel_path
            )
            
            if not optimization_result['success']:
                return {
                    'success': False,
                    'error': f"Template optimization failed: {optimization_result['error']}"
                }
            
            # Generate output path if not provided
            if not output_path:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                template_name = os.path.splitext(os.path.basename(template_path))[0]
                filename = f"report_{template_name}_{timestamp}.tex"
                output_dir = os.path.join(os.path.dirname(__file__), '../../static/generated')
                os.makedirs(output_dir, exist_ok=True)
                output_path = os.path.join(output_dir, filename)
            
            # Render template
            from .template_converter import TemplateConverter
            
            # Convert template if needed
            syntax_type = TemplateConverter.analyze_template_syntax(template_content)
            if syntax_type in ['mustache', 'mixed']:
                template_content = TemplateConverter.mustache_to_jinja2(template_content)
            
            # Render with context
            context = optimization_result['enhanced_context']
            template = Template(template_content)
            rendered_content = template.render(context)
            
            # Save rendered report
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(rendered_content)
            
            result = {
                'success': True,
                'report_path': output_path,
                'filename': os.path.basename(output_path),
                'template_used': template_path,
                'data_source': excel_path,
                'generation_timestamp': datetime.now().isoformat(),
                'context_summary': {
                    'total_fields': len([k for k in context.keys() if not k.startswith('_')]),
                    'data_rows': extracted_data.get('total_rows', 0),
                    'sheets_processed': extracted_data.get('sheets_processed', 0)
                }
            }
            
            logger.info(f"Successfully generated report: {output_path}")
            return result
            
        except Exception as e:
            logger.error(f"Error generating report from Excel: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def create_automated_workflow(self, template_path: str, workflow_name: str) -> Dict[str, Any]:
        """
        Create a complete automated workflow: Template → Form → Excel → Report
        
        Args:
            template_path: Path to the template file
            workflow_name: Name for the workflow
            
        Returns:
            Dict with workflow configuration
        """
        try:
            logger.info(f"Creating automated workflow: {workflow_name}")
            
            # Step 1: Create form from template
            form_result = self.create_form_from_template(template_path, f"{workflow_name} - Data Collection")
            
            if not form_result['success']:
                return {
                    'success': False,
                    'error': f"Failed to create form: {form_result['error']}"
                }
            
            # Step 2: Save the form to database (this would be called separately)
            # For now, return the workflow configuration
            
            workflow_config = {
                'success': True,
                'workflow_name': workflow_name,
                'template_path': template_path,
                'form_schema': form_result['form_schema'],
                'automation_steps': [
                    {
                        'step': 1,
                        'name': 'Form Creation',
                        'description': 'Dynamic form created from template placeholders',
                        'status': 'completed',
                        'fields_detected': form_result['detected_fields']
                    },
                    {
                        'step': 2,
                        'name': 'Data Collection',
                        'description': 'Collect data through generated form',
                        'status': 'ready',
                        'action': 'Deploy form and collect submissions'
                    },
                    {
                        'step': 3,
                        'name': 'Excel Export',
                        'description': 'Export collected data to formatted Excel file',
                        'status': 'pending',
                        'action': 'Call export_form_data_to_excel() when ready'
                    },
                    {
                        'step': 4,
                        'name': 'Report Generation',
                        'description': 'Generate final report from Excel data',
                        'status': 'pending',
                        'action': 'Call generate_report_from_excel() when ready'
                    }
                ],
                'estimated_completion_time': '5-10 minutes per submission batch',
                'automation_features': [
                    'Smart field type detection',
                    'Professional Excel formatting',
                    'Comprehensive analytics',
                    'Template-driven report generation',
                    'QR code support for easy form access'
                ]
            }
            
            logger.info(f"Successfully created automated workflow: {workflow_name}")
            return workflow_config
            
        except Exception as e:
            logger.error(f"Error creating automated workflow: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
