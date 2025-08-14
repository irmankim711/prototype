"""
Enhanced Excel Data Parser for Report Builder
Provides smart data analysis and field detection for better report generation
"""

import pandas as pd
import numpy as np
import os
import json
import logging
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import openpyxl
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

class EnhancedExcelParser:
    def __init__(self):
        """Initialize Enhanced Excel Parser"""
        self.supported_formats = ['.xlsx', '.xls', '.csv']
        
    def parse_excel_file(self, file_path: str) -> Dict[str, Any]:
        """
        Parse Excel file and extract comprehensive data analysis
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dict containing parsed data and analysis
        """
        try:
            if not os.path.exists(file_path):
                return {"success": False, "error": "File not found"}
            
            file_ext = os.path.splitext(file_path)[1].lower()
            if file_ext not in self.supported_formats:
                return {"success": False, "error": f"Unsupported file format: {file_ext}"}
            
            # Parse file based on extension
            if file_ext == '.csv':
                return self._parse_csv_file(file_path)
            else:
                return self._parse_excel_file(file_path)
                
        except Exception as e:
            logger.error(f"Error parsing Excel file: {str(e)}")
            return {"success": False, "error": str(e)}
    
    def _parse_excel_file(self, file_path: str) -> Dict[str, Any]:
        """Parse Excel file (.xlsx, .xls)"""
        try:
            # Load workbook to get sheet information
            workbook = load_workbook(file_path, data_only=True)
            sheet_names = workbook.sheetnames
            
            # Parse all sheets
            sheets_data = {}
            for sheet_name in sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
                    sheets_data[sheet_name] = self._analyze_dataframe(df, sheet_name)
                except Exception as e:
                    logger.warning(f"Failed to parse sheet '{sheet_name}': {str(e)}")
                    continue
            
            # Get primary sheet (first sheet with data)
            primary_sheet = None
            for sheet_name, sheet_data in sheets_data.items():
                if sheet_data['success'] and sheet_data['total_rows'] > 0:
                    primary_sheet = sheet_data
                    break
            
            if not primary_sheet:
                return {"success": False, "error": "No valid data found in any sheet"}
            
            return {
                "success": True,
                "file_path": file_path,
                "sheets": sheets_data,
                "primary_sheet": primary_sheet,
                "sheet_count": len(sheet_names),
                "analysis_timestamp": datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Error parsing Excel file: {str(e)}")
            return {"success": False, "error": str(e)}
    
    def _parse_csv_file(self, file_path: str) -> Dict[str, Any]:
        """Parse CSV file"""
        try:
            df = pd.read_csv(file_path)
            sheet_data = self._analyze_dataframe(df, "Sheet1")
            
            return {
                "success": True,
                "file_path": file_path,
                "sheets": {"Sheet1": sheet_data},
                "primary_sheet": sheet_data,
                "sheet_count": 1,
                "analysis_timestamp": datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Error parsing CSV file: {str(e)}")
            return {"success": False, "error": str(e)}
    
    def _analyze_dataframe(self, df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """
        Analyze DataFrame and extract comprehensive information
        
        Args:
            df: Pandas DataFrame
            sheet_name: Name of the sheet
            
        Returns:
            Dict containing analysis results
        """
        try:
            if df.empty:
                return {"success": False, "error": "Empty dataframe"}
            
            # Basic information
            total_rows, total_cols = df.shape
            headers = df.columns.tolist()
            
            # Clean and prepare data
            df_clean = df.dropna(how='all')  # Remove completely empty rows
            df_clean = df_clean.dropna(axis=1, how='all')  # Remove completely empty columns
            
            # Data type analysis
            data_types = self._analyze_data_types(df_clean)
            
            # Field categorization
            field_categories = self._categorize_fields(df_clean, headers)
            
            # Statistical analysis
            statistics = self._generate_statistics(df_clean)
            
            # Sample data (first 10 rows)
            sample_data = self._get_sample_data(df_clean, 10)
            
            # Data quality assessment
            quality_assessment = self._assess_data_quality(df_clean)
            
            # Potential relationships
            relationships = self._detect_relationships(df_clean)
            
            return {
                "success": True,
                "sheet_name": sheet_name,
                "total_rows": total_rows,
                "total_cols": total_cols,
                "clean_rows": len(df_clean),
                "clean_cols": len(df_clean.columns),
                "headers": headers,
                "data_types": data_types,
                "field_categories": field_categories,
                "statistics": statistics,
                "sample_data": sample_data,
                "quality_assessment": quality_assessment,
                "relationships": relationships,
                "summary": self._generate_data_summary(df_clean, headers)
            }
            
        except Exception as e:
            logger.error(f"Error analyzing dataframe: {str(e)}")
            return {"success": False, "error": str(e)}
    
    def _analyze_data_types(self, df: pd.DataFrame) -> Dict[str, str]:
        """Analyze and categorize data types"""
        data_types = {}
        
        for column in df.columns:
            series = df[column].dropna()
            if len(series) == 0:
                data_types[column] = "empty"
                continue
            
            # Check for numeric types
            if pd.api.types.is_numeric_dtype(series):
                if series.dtype == 'int64' or series.dtype == 'int32':
                    data_types[column] = "integer"
                else:
                    data_types[column] = "decimal"
            # Check for datetime types
            elif pd.api.types.is_datetime64_any_dtype(series):
                data_types[column] = "datetime"
            else:
                # Try to detect specific patterns
                sample_values = series.astype(str).head(20).tolist()
                detected_type = self._detect_field_type(sample_values)
                data_types[column] = detected_type
        
        return data_types
    
    def _detect_field_type(self, sample_values: List[str]) -> str:
        """Detect field type based on sample values"""
        # Date patterns
        date_patterns = 0
        email_patterns = 0
        phone_patterns = 0
        currency_patterns = 0
        percentage_patterns = 0
        
        for value in sample_values:
            value_str = str(value).strip().lower()
            
            # Date detection
            if any(char in value_str for char in ['/', '-', ':', 'jan', 'feb', 'mar', 'apr', 'may', 'jun']):
                date_patterns += 1
            
            # Email detection
            if '@' in value_str and '.' in value_str:
                email_patterns += 1
            
            # Phone detection
            if any(char in value_str for char in ['+', '(', ')', '-']) and any(c.isdigit() for c in value_str):
                phone_patterns += 1
            
            # Currency detection
            if any(symbol in value_str for symbol in ['$', '€', '£', 'usd', 'eur', 'gbp']):
                currency_patterns += 1
            
            # Percentage detection
            if '%' in value_str:
                percentage_patterns += 1
        
        total_samples = len(sample_values)
        if total_samples == 0:
            return "text"
        
        # Determine type based on patterns
        if date_patterns / total_samples > 0.5:
            return "date"
        elif email_patterns / total_samples > 0.5:
            return "email"
        elif phone_patterns / total_samples > 0.5:
            return "phone"
        elif currency_patterns / total_samples > 0.5:
            return "currency"
        elif percentage_patterns / total_samples > 0.5:
            return "percentage"
        else:
            return "text"
    
    def _categorize_fields(self, df: pd.DataFrame, headers: List[str]) -> Dict[str, str]:
        """Categorize fields based on headers and content"""
        categories = {}
        
        for header in headers:
            header_lower = header.lower().strip()
            
            # Personal information
            if any(keyword in header_lower for keyword in ['name', 'nama', 'participant', 'student', 'employee']):
                categories[header] = "personal_info"
            elif any(keyword in header_lower for keyword in ['email', 'e-mail', 'mail']):
                categories[header] = "contact_info"
            elif any(keyword in header_lower for keyword in ['phone', 'tel', 'mobile', 'contact']):
                categories[header] = "contact_info"
            elif any(keyword in header_lower for keyword in ['address', 'location', 'city', 'state']):
                categories[header] = "location_info"
            
            # Identifiers
            elif any(keyword in header_lower for keyword in ['id', 'ic', 'nric', 'passport', 'number', 'no']):
                categories[header] = "identifier"
            
            # Scores and metrics
            elif any(keyword in header_lower for keyword in ['score', 'mark', 'grade', 'point', 'rating']):
                categories[header] = "performance_metric"
            elif any(keyword in header_lower for keyword in ['pre', 'post', 'before', 'after']):
                categories[header] = "performance_metric"
            
            # Dates and time
            elif any(keyword in header_lower for keyword in ['date', 'time', 'timestamp', 'created', 'updated']):
                categories[header] = "temporal_info"
            
            # Financial
            elif any(keyword in header_lower for keyword in ['salary', 'wage', 'cost', 'price', 'amount', 'fee']):
                categories[header] = "financial_info"
            
            # Status and categories
            elif any(keyword in header_lower for keyword in ['status', 'state', 'category', 'type', 'department']):
                categories[header] = "categorical_info"
            
            else:
                categories[header] = "general_info"
        
        return categories
    
    def _generate_statistics(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Generate statistical analysis of the data"""
        stats = {}
        
        numeric_columns = df.select_dtypes(include=[np.number]).columns.tolist()
        text_columns = df.select_dtypes(include=['object']).columns.tolist()
        
        # Numeric statistics
        if numeric_columns:
            numeric_stats = {}
            for col in numeric_columns:
                series = df[col].dropna()
                if len(series) > 0:
                    numeric_stats[col] = {
                        "mean": float(series.mean()),
                        "median": float(series.median()),
                        "min": float(series.min()),
                        "max": float(series.max()),
                        "std": float(series.std()) if len(series) > 1 else 0,
                        "count": int(series.count())
                    }
            stats["numeric"] = numeric_stats
        
        # Text statistics
        if text_columns:
            text_stats = {}
            for col in text_columns:
                series = df[col].dropna().astype(str)
                if len(series) > 0:
                    value_counts = series.value_counts().head(10)
                    text_stats[col] = {
                        "unique_count": int(series.nunique()),
                        "most_common": value_counts.to_dict(),
                        "avg_length": float(series.str.len().mean()),
                        "count": int(series.count())
                    }
            stats["text"] = text_stats
        
        return stats
    
    def _get_sample_data(self, df: pd.DataFrame, sample_size: int = 10) -> List[Dict[str, Any]]:
        """Get sample data rows"""
        try:
            sample_df = df.head(sample_size)
            # Convert to records and handle NaN values
            records = sample_df.to_dict('records')
            
            # Clean NaN values
            cleaned_records = []
            for record in records:
                cleaned_record = {}
                for key, value in record.items():
                    if pd.isna(value):
                        cleaned_record[key] = None
                    elif isinstance(value, (np.integer, np.floating)):
                        cleaned_record[key] = float(value) if not np.isnan(value) else None
                    else:
                        cleaned_record[key] = str(value) if value is not None else None
                cleaned_records.append(cleaned_record)
            
            return cleaned_records
        except Exception as e:
            logger.error(f"Error getting sample data: {str(e)}")
            return []
    
    def _assess_data_quality(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Assess data quality"""
        total_cells = df.shape[0] * df.shape[1]
        missing_cells = df.isnull().sum().sum()
        
        quality = {
            "completeness": {
                "total_cells": int(total_cells),
                "missing_cells": int(missing_cells),
                "completeness_percentage": float((total_cells - missing_cells) / total_cells * 100) if total_cells > 0 else 0
            },
            "consistency": {},
            "uniqueness": {}
        }
        
        # Check for duplicate rows
        duplicate_rows = df.duplicated().sum()
        quality["uniqueness"]["duplicate_rows"] = int(duplicate_rows)
        quality["uniqueness"]["unique_rows"] = int(len(df) - duplicate_rows)
        
        # Column-wise quality assessment
        column_quality = {}
        for col in df.columns:
            series = df[col]
            column_quality[col] = {
                "missing_count": int(series.isnull().sum()),
                "missing_percentage": float(series.isnull().sum() / len(series) * 100),
                "unique_values": int(series.nunique()),
                "duplicate_values": int(len(series) - series.nunique())
            }
        
        quality["column_quality"] = column_quality
        
        return quality
    
    def _detect_relationships(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Detect potential relationships between columns"""
        relationships = {
            "correlations": {},
            "potential_keys": [],
            "grouped_fields": {}
        }
        
        try:
            # Numeric correlations
            numeric_df = df.select_dtypes(include=[np.number])
            if len(numeric_df.columns) > 1:
                corr_matrix = numeric_df.corr()
                # Convert to regular dict for JSON serialization
                correlations = {}
                for i, col1 in enumerate(corr_matrix.columns):
                    for j, col2 in enumerate(corr_matrix.columns):
                        if i < j:  # Avoid duplicates
                            corr_value = corr_matrix.iloc[i, j]
                            if not pd.isna(corr_value) and abs(corr_value) > 0.5:
                                correlations[f"{col1}_vs_{col2}"] = float(corr_value)
                relationships["correlations"] = correlations
            
            # Potential primary keys (unique values)
            for col in df.columns:
                series = df[col].dropna()
                if len(series) > 0 and series.nunique() == len(series):
                    relationships["potential_keys"].append(col)
            
            # Group related fields
            field_groups = {}
            for col in df.columns:
                col_lower = col.lower()
                if 'pre' in col_lower or 'before' in col_lower:
                    if 'pre_post_scores' not in field_groups:
                        field_groups['pre_post_scores'] = []
                    field_groups['pre_post_scores'].append(col)
                elif 'post' in col_lower or 'after' in col_lower:
                    if 'pre_post_scores' not in field_groups:
                        field_groups['pre_post_scores'] = []
                    field_groups['pre_post_scores'].append(col)
                elif any(keyword in col_lower for keyword in ['name', 'nama']):
                    if 'personal_info' not in field_groups:
                        field_groups['personal_info'] = []
                    field_groups['personal_info'].append(col)
                elif any(keyword in col_lower for keyword in ['email', 'phone', 'contact']):
                    if 'contact_info' not in field_groups:
                        field_groups['contact_info'] = []
                    field_groups['contact_info'].append(col)
            
            relationships["grouped_fields"] = field_groups
            
        except Exception as e:
            logger.error(f"Error detecting relationships: {str(e)}")
        
        return relationships
    
    def _generate_data_summary(self, df: pd.DataFrame, headers: List[str]) -> Dict[str, Any]:
        """Generate high-level data summary"""
        return {
            "total_records": len(df),
            "total_fields": len(headers),
            "data_types_summary": {
                "numeric_fields": len(df.select_dtypes(include=[np.number]).columns),
                "text_fields": len(df.select_dtypes(include=['object']).columns),
                "datetime_fields": len(df.select_dtypes(include=['datetime']).columns)
            },
            "key_characteristics": {
                "has_personal_data": any(keyword in ' '.join(headers).lower() for keyword in ['name', 'email', 'phone']),
                "has_performance_data": any(keyword in ' '.join(headers).lower() for keyword in ['score', 'grade', 'mark', 'rating']),
                "has_temporal_data": any(keyword in ' '.join(headers).lower() for keyword in ['date', 'time', 'created']),
                "has_financial_data": any(keyword in ' '.join(headers).lower() for keyword in ['salary', 'cost', 'price', 'amount'])
            }
        }

# Global instance
enhanced_excel_parser = EnhancedExcelParser()
