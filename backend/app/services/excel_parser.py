"""
Excel parsing service for automatic data block detection and conversion.
Handles Excel files (.xlsx, .xls) and Google Sheets exports.
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.cell import Cell
from typing import List, Dict, Any, Tuple, Optional, Union
import re
from datetime import datetime
import logging
from io import BytesIO
import xlrd

logger = logging.getLogger(__name__)


class ExcelTableDetector:
    """Detects and extracts data tables from Excel sheets."""
    
    def __init__(self, file_path: Union[str, BytesIO], file_extension: str = None):
        self.file_path = file_path
        self.file_extension = file_extension or self._detect_file_type()
        self.workbook = None
        self.detected_tables = []
        
    def _detect_file_type(self) -> str:
        """Detect file type from file path or content."""
        if isinstance(self.file_path, str):
            if self.file_path.lower().endswith('.xlsx'):
                return 'xlsx'
            elif self.file_path.lower().endswith('.xls'):
                return 'xls'
        return 'xlsx'  # Default to xlsx
    
    def detect_all_tables(self) -> List[Dict[str, Any]]:
        """
        Detect all data tables across all sheets.
        Returns a list of table metadata with data.
        """
        try:
            if self.file_extension == 'xlsx':
                return self._detect_tables_openpyxl()
            else:
                return self._detect_tables_xlrd()
        except Exception as e:
            logger.error(f"Error detecting tables: {str(e)}")
            raise
    
    def _detect_tables_openpyxl(self) -> List[Dict[str, Any]]:
        """Detect tables using openpyxl for .xlsx files."""
        tables = []
        self.workbook = load_workbook(self.file_path, data_only=True)
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet_tables = self._find_tables_in_sheet_openpyxl(sheet, sheet_name)
            tables.extend(sheet_tables)
        
        return tables
    
    def _detect_tables_xlrd(self) -> List[Dict[str, Any]]:
        """Detect tables using xlrd for .xls files."""
        tables = []
        workbook = xlrd.open_workbook(self.file_path)
        
        for sheet_idx in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_idx)
            sheet_name = workbook.sheet_names()[sheet_idx]
            sheet_tables = self._find_tables_in_sheet_xlrd(sheet, sheet_name)
            tables.extend(sheet_tables)
        
        return tables
    
    def _find_tables_in_sheet_openpyxl(self, sheet, sheet_name: str) -> List[Dict[str, Any]]:
        """Find all tables in a single sheet using openpyxl."""
        tables = []
        
        # Get all non-empty cells
        non_empty_cells = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip():
                    non_empty_cells.append((cell.row - 1, cell.column - 1))  # Convert to 0-based
        
        if not non_empty_cells:
            return tables
        
        # Group cells into contiguous blocks
        table_blocks = self._group_cells_into_blocks(non_empty_cells)
        
        # Process each block
        for block_idx, block in enumerate(table_blocks):
            table_data = self._extract_table_data_openpyxl(sheet, block)
            if table_data and len(table_data) > 1:  # At least header + 1 row
                table_name = f"{sheet_name}_Table{block_idx + 1}" if len(table_blocks) > 1 else sheet_name
                
                tables.append({
                    'name': table_name,
                    'sheet_name': sheet_name,
                    'data': table_data,
                    'headers': table_data[0] if table_data else [],
                    'row_count': len(table_data) - 1 if table_data else 0,
                    'column_count': len(table_data[0]) if table_data else 0,
                    'range': self._get_table_range(block),
                    'data_types': self._infer_column_types(table_data)
                })
        
        return tables
    
    def _find_tables_in_sheet_xlrd(self, sheet, sheet_name: str) -> List[Dict[str, Any]]:
        """Find all tables in a single sheet using xlrd."""
        tables = []
        
        # Get all non-empty cells
        non_empty_cells = []
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                cell_value = sheet.cell_value(row_idx, col_idx)
                if cell_value and str(cell_value).strip():
                    non_empty_cells.append((row_idx, col_idx))
        
        if not non_empty_cells:
            return tables
        
        # Group cells into contiguous blocks
        table_blocks = self._group_cells_into_blocks(non_empty_cells)
        
        # Process each block
        for block_idx, block in enumerate(table_blocks):
            table_data = self._extract_table_data_xlrd(sheet, block)
            if table_data and len(table_data) > 1:  # At least header + 1 row
                table_name = f"{sheet_name}_Table{block_idx + 1}" if len(table_blocks) > 1 else sheet_name
                
                tables.append({
                    'name': table_name,
                    'sheet_name': sheet_name,
                    'data': table_data,
                    'headers': table_data[0] if table_data else [],
                    'row_count': len(table_data) - 1 if table_data else 0,
                    'column_count': len(table_data[0]) if table_data else 0,
                    'range': self._get_table_range(block),
                    'data_types': self._infer_column_types(table_data)
                })
        
        return tables
    
    def _group_cells_into_blocks(self, cells: List[Tuple[int, int]]) -> List[List[Tuple[int, int]]]:
        """Group adjacent cells into contiguous rectangular blocks."""
        if not cells:
            return []
        
        # Sort cells by row, then column
        sorted_cells = sorted(cells)
        
        # Find contiguous blocks using flood fill algorithm
        visited = set()
        blocks = []
        
        for cell in sorted_cells:
            if cell in visited:
                continue
            
            # Start a new block
            block = []
            stack = [cell]
            
            while stack:
                current = stack.pop()
                if current in visited:
                    continue
                
                visited.add(current)
                block.append(current)
                
                # Check adjacent cells (4-directional)
                row, col = current
                for dr, dc in [(0, 1), (0, -1), (1, 0), (-1, 0)]:
                    adj_cell = (row + dr, col + dc)
                    if adj_cell in sorted_cells and adj_cell not in visited:
                        stack.append(adj_cell)
            
            if len(block) >= 4:  # Minimum table size (2x2)
                blocks.append(block)
        
        return blocks
    
    def _extract_table_data_openpyxl(self, sheet, block: List[Tuple[int, int]]) -> List[List[Any]]:
        """Extract data from a table block using openpyxl."""
        if not block:
            return []
        
        # Find table boundaries
        min_row = min(cell[0] for cell in block)
        max_row = max(cell[0] for cell in block)
        min_col = min(cell[1] for cell in block)
        max_col = max(cell[1] for cell in block)
        
        # Extract data
        table_data = []
        for row_idx in range(min_row, max_row + 1):
            row_data = []
            for col_idx in range(min_col, max_col + 1):
                cell = sheet.cell(row=row_idx + 1, column=col_idx + 1)  # Convert to 1-based
                value = self._clean_cell_value(cell.value)
                row_data.append(value)
            table_data.append(row_data)
        
        return self._clean_table_data(table_data)
    
    def _extract_table_data_xlrd(self, sheet, block: List[Tuple[int, int]]) -> List[List[Any]]:
        """Extract data from a table block using xlrd."""
        if not block:
            return []
        
        # Find table boundaries
        min_row = min(cell[0] for cell in block)
        max_row = max(cell[0] for cell in block)
        min_col = min(cell[1] for cell in block)
        max_col = max(cell[1] for cell in block)
        
        # Extract data
        table_data = []
        for row_idx in range(min_row, max_row + 1):
            row_data = []
            for col_idx in range(min_col, max_col + 1):
                if row_idx < sheet.nrows and col_idx < sheet.ncols:
                    value = sheet.cell_value(row_idx, col_idx)
                    value = self._clean_cell_value(value)
                else:
                    value = None
                row_data.append(value)
            table_data.append(row_data)
        
        return self._clean_table_data(table_data)
    
    def _clean_cell_value(self, value: Any) -> Any:
        """Clean and normalize cell value."""
        if value is None:
            return None
        
        # Handle strings
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            
            # Try to convert to number
            try:
                if '.' in value or 'e' in value.lower():
                    return float(value)
                else:
                    return int(value)
            except ValueError:
                pass
            
            # Try to parse date
            for date_format in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S']:
                try:
                    return datetime.strptime(value, date_format).isoformat()
                except ValueError:
                    continue
            
            return value
        
        # Handle numbers
        if isinstance(value, (int, float)):
            return value
        
        # Handle datetime
        if hasattr(value, 'isoformat'):
            return value.isoformat()
        
        return str(value) if value is not None else None
    
    def _clean_table_data(self, table_data: List[List[Any]]) -> List[List[Any]]:
        """Clean table data by removing empty rows/columns and normalizing."""
        if not table_data:
            return []
        
        # Remove completely empty rows
        cleaned_data = []
        for row in table_data:
            if any(cell is not None and str(cell).strip() for cell in row):
                cleaned_data.append(row)
        
        if not cleaned_data:
            return []
        
        # Ensure all rows have the same length
        max_cols = max(len(row) for row in cleaned_data)
        for row in cleaned_data:
            while len(row) < max_cols:
                row.append(None)
        
        # Generate column headers if first row doesn't look like headers
        first_row = cleaned_data[0]
        if not self._looks_like_headers(first_row):
            headers = [f"Col{i+1}" for i in range(len(first_row))]
            cleaned_data.insert(0, headers)
        
        return cleaned_data
    
    def _looks_like_headers(self, row: List[Any]) -> bool:
        """Determine if a row looks like column headers."""
        if not row:
            return False
        
        # Check if all cells are strings and not numbers
        for cell in row:
            if cell is None:
                continue
            if isinstance(cell, (int, float)):
                return False
            if isinstance(cell, str) and cell.isdigit():
                return False
        
        return True
    
    def _get_table_range(self, block: List[Tuple[int, int]]) -> str:
        """Get Excel-style range for a table block."""
        if not block:
            return ""
        
        min_row = min(cell[0] for cell in block) + 1  # Convert to 1-based
        max_row = max(cell[0] for cell in block) + 1
        min_col = min(cell[1] for cell in block)
        max_col = max(cell[1] for cell in block)
        
        def col_to_letter(col_num):
            result = ""
            while col_num >= 0:
                result = chr(col_num % 26 + ord('A')) + result
                col_num = col_num // 26 - 1
            return result
        
        start_cell = f"{col_to_letter(min_col)}{min_row}"
        end_cell = f"{col_to_letter(max_col)}{max_row}"
        
        return f"{start_cell}:{end_cell}"
    
    def _infer_column_types(self, table_data: List[List[Any]]) -> List[str]:
        """Infer data types for each column."""
        if len(table_data) < 2:  # Need at least header + 1 data row
            return []
        
        data_rows = table_data[1:]  # Skip header
        column_count = len(table_data[0])
        column_types = []
        
        for col_idx in range(column_count):
            column_values = [row[col_idx] for row in data_rows if col_idx < len(row)]
            non_null_values = [v for v in column_values if v is not None]
            
            if not non_null_values:
                column_types.append('text')
                continue
            
            # Check for numeric types
            numeric_count = sum(1 for v in non_null_values if isinstance(v, (int, float)))
            if numeric_count / len(non_null_values) > 0.8:
                if all(isinstance(v, int) or (isinstance(v, float) and v.is_integer()) 
                       for v in non_null_values if isinstance(v, (int, float))):
                    column_types.append('integer')
                else:
                    column_types.append('float')
                continue
            
            # Check for date types
            date_count = sum(1 for v in non_null_values 
                           if isinstance(v, str) and self._is_date_string(v))
            if date_count / len(non_null_values) > 0.8:
                column_types.append('date')
                continue
            
            # Default to text
            column_types.append('text')
        
        return column_types
    
    def _is_date_string(self, value: str) -> bool:
        """Check if a string looks like a date."""
        date_patterns = [
            r'\d{4}-\d{2}-\d{2}',
            r'\d{2}/\d{2}/\d{4}',
            r'\d{1,2}/\d{1,2}/\d{4}',
            r'\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}'
        ]
        
        for pattern in date_patterns:
            if re.match(pattern, value):
                return True
        
        return False


class ExcelParserService:
    """Main service for parsing Excel files and converting to structured data."""
    
    def __init__(self):
        self.supported_extensions = ['.xlsx', '.xls']
        self.max_file_size = 50 * 1024 * 1024  # 50MB
    
    def parse_excel_file(self, file_path: Union[str, BytesIO], 
                        filename: str = None) -> Dict[str, Any]:
        """
        Parse an Excel file and return structured data.
        
        Args:
            file_path: Path to file or BytesIO object
            filename: Original filename (for extension detection)
        
        Returns:
            Dictionary containing parsed tables and metadata
        """
        try:
            # Detect file extension
            file_extension = None
            if filename:
                file_extension = self._get_file_extension(filename)
            
            # Initialize detector
            detector = ExcelTableDetector(file_path, file_extension)
            
            # Detect all tables
            tables = detector.detect_all_tables()
            
            # Generate response
            response = {
                'success': True,
                'filename': filename or 'unknown',
                'tables_count': len(tables),
                'tables': tables,
                'metadata': {
                    'parsed_at': datetime.utcnow().isoformat(),
                    'total_rows': sum(table['row_count'] for table in tables),
                    'total_columns': sum(table['column_count'] for table in tables),
                    'sheets_processed': len(set(table['sheet_name'] for table in tables))
                }
            }
            
            return response
            
        except Exception as e:
            logger.error(f"Error parsing Excel file: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'filename': filename or 'unknown'
            }
    
    def _get_file_extension(self, filename: str) -> str:
        """Extract file extension from filename."""
        if '.' in filename:
            return filename.split('.')[-1].lower()
        return 'xlsx'  # Default
    
    def validate_file(self, file_size: int, filename: str) -> Tuple[bool, str]:
        """
        Validate file before processing.
        
        Returns:
            Tuple of (is_valid, error_message)
        """
        # Check file size
        if file_size > self.max_file_size:
            return False, f"File size ({file_size / 1024 / 1024:.1f}MB) exceeds maximum allowed size (50MB)"
        
        # Check file extension
        if filename:
            extension = '.' + self._get_file_extension(filename)
            if extension not in self.supported_extensions:
                return False, f"Unsupported file type. Supported types: {', '.join(self.supported_extensions)}"
        
        return True, ""
    
    def convert_to_csv_format(self, table_data: List[List[Any]]) -> str:
        """Convert table data to CSV format."""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        for row in table_data:
            # Convert all values to strings, handling None values
            csv_row = [str(cell) if cell is not None else '' for cell in row]
            writer.writerow(csv_row)
        
        return output.getvalue()
    
    def generate_sql_schema(self, table: Dict[str, Any]) -> str:
        """Generate SQL CREATE TABLE statement for a table."""
        table_name = table['name'].replace(' ', '_').replace('-', '_')
        headers = table['headers']
        data_types = table['data_types']
        
        sql_parts = [f"CREATE TABLE {table_name} ("]
        
        for i, (header, data_type) in enumerate(zip(headers, data_types)):
            column_name = str(header).replace(' ', '_').replace('-', '_') if header else f"col_{i+1}"
            
            # Map data types to SQL types
            sql_type_map = {
                'integer': 'INTEGER',
                'float': 'DECIMAL(10,2)',
                'date': 'DATE',
                'text': 'TEXT'
            }
            
            sql_type = sql_type_map.get(data_type, 'TEXT')
            sql_parts.append(f"    {column_name} {sql_type}")
            
            if i < len(headers) - 1:
                sql_parts[-1] += ","
        
        sql_parts.append(");")
        
        return "\n".join(sql_parts)
