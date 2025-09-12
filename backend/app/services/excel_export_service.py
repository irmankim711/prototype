"""
Enhanced Excel Export Service
Handles export of form data to Excel with production-ready features
"""

import os
import json
import uuid
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
import logging

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

from .. import db
from ..models import Form, FormSubmission, User
from ..core.exceptions import ReportGenerationError as ExportError

logger = logging.getLogger(__name__)

class ExcelExportService:
    """Service for exporting form data to Excel with enhanced formatting"""
    
    def __init__(self, upload_folder: str = None):
        self.upload_folder = upload_folder or os.path.join(os.getcwd(), 'uploads', 'exports')
        self.ensure_upload_directory()
    
    def ensure_upload_directory(self):
        """Ensure the upload directory exists"""
        Path(self.upload_folder).mkdir(parents=True, exist_ok=True)
    
    def generate_timestamp_filename(self, prefix: str, form_title: str = None) -> str:
        """Generate a unique filename with timestamp"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        unique_id = str(uuid.uuid4())[:8]
        
        if form_title:
            clean_title = "".join(c for c in form_title if c.isalnum() or c in (' ', '-', '_')).rstrip()
            clean_title = clean_title.replace(' ', '_')[:30]
            return f"{prefix}_{clean_title}_{timestamp}_{unique_id}.xlsx"
        else:
            return f"{prefix}_{timestamp}_{unique_id}.xlsx"
    
    def export_form_submissions(self, form_id: int, user_id: int, options: Dict[str, Any] = None) -> Tuple[str, int]:
        """Export form submissions to Excel"""
        try:
            form = Form.query.get(form_id)
            if not form:
                raise ExportError(f"Form {form_id} not found")
            
            options = options or {}
            date_range = options.get('date_range', {})
            max_records = options.get('max_records', 10000)
            
            # Build query
            query = FormSubmission.query.filter_by(form_id=form_id)
            
            # Apply date filters
            if date_range.get('start'):
                start_date = datetime.fromisoformat(date_range['start'])
                query = query.filter(FormSubmission.submitted_at >= start_date)
            
            if date_range.get('end'):
                end_date = datetime.fromisoformat(date_range['end'])
                query = query.filter(FormSubmission.submitted_at <= end_date)
            
            query = query.limit(max_records)
            submissions = query.all()
            
            # Generate filename and create workbook
            filename = self.generate_timestamp_filename("export", form.title)
            file_path = os.path.join(self.upload_folder, filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Form Submissions"
            
            # Export data
            self._export_submissions_to_sheet(ws, submissions, form)
            
            # Save and return
            wb.save(file_path)
            file_size = os.path.getsize(file_path)
            
            logger.info(f"Excel export completed for form {form_id}: {file_path} ({file_size} bytes)")
            return file_path, file_size
            
        except Exception as e:
            logger.error(f"Error exporting form {form_id} to Excel: {str(e)}")
            raise ExportError(f"Failed to export form to Excel: {str(e)}")
    
    def _export_submissions_to_sheet(self, ws, submissions, form):
        """Export submissions data to Excel worksheet"""
        # Get form schema
        form_schema = form.schema or {}
        fields = form_schema.get('fields', [])
        
        # Prepare headers
        headers = ['ID', 'Submitted At', 'Status', 'Submitter Email']
        for field in fields:
            field_label = field.get('label', field.get('name', 'Unknown'))
            headers.append(field_label)
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write data rows
        for row_idx, submission in enumerate(submissions, 2):
            col = 1
            
            ws.cell(row=row_idx, column=col, value=submission.id)
            col += 1
            
            ws.cell(row=row_idx, column=col, value=submission.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if submission.submitted_at else '')
            col += 1
            
            ws.cell(row=row_idx, column=col, value=submission.status)
            col += 1
            
            ws.cell(row=row_idx, column=col, value=submission.submitter_email or '')
            col += 1
            
            # Form field data
            submission_data = submission.data or {}
            for field in fields:
                field_name = field.get('name', 'Unknown')
                field_value = submission_data.get(field_name, '')
                
                if isinstance(field_value, (dict, list)):
                    field_value = json.dumps(field_value, ensure_ascii=False)
                
                ws.cell(row=row_idx, column=col, value=str(field_value))
                col += 1
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

# Global instance
excel_export_service = ExcelExportService()
