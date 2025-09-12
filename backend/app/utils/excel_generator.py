"""
Excel generation utility using pandas and openpyxl
"""

import os
import pandas as pd
from datetime import datetime
from typing import Dict, List, Tuple, Any, Optional
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import uuid

logger = logging.getLogger(__name__)

class ExcelGenerator:
    """Generate Excel files from form submission data"""
    
    def __init__(self, export_config, submissions):
        self.export_config = export_config
        self.submissions = submissions
        self.template_config = export_config.template_config or {}
        
        # Create reports directory if it doesn't exist
        self.reports_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'reports')
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def generate(self) -> Tuple[str, str, int]:
        """
        Generate Excel file and return (file_path, file_name, file_size)
        """
        try:
            # Generate unique filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name = f"{self.export_config.name}_{timestamp}.xlsx"
            file_path = os.path.join(self.reports_dir, file_name)
            
            # Convert submissions to DataFrame
            df = self._create_dataframe()
            
            # Generate Excel file based on template configuration
            if self.template_config.get('use_advanced_template'):
                self._create_advanced_excel(df, file_path)
            else:
                self._create_basic_excel(df, file_path)
            
            # Get file size
            file_size = os.path.getsize(file_path)
            
            logger.info(f"Excel file generated: {file_path} ({file_size} bytes)")
            
            return file_path, file_name, file_size
            
        except Exception as e:
            logger.error(f"Excel generation failed: {str(e)}")
            raise e
    
    def _create_dataframe(self) -> pd.DataFrame:
        """Convert form submissions to pandas DataFrame"""
        try:
            # Collect all data
            data_rows = []
            
            for submission in self.submissions:
                row = {
                    'Submission ID': submission.id,
                    'External ID': submission.external_id,
                    'Data Source': submission.data_source.name,
                    'Source Type': submission.data_source.source_type,
                    'Submitted At': submission.submitted_at,
                    'Submitter Email': submission.submitter_email,
                    'Submitter Name': submission.submitter_name,
                    'Processing Status': submission.processing_status
                }
                
                # Add normalized data fields
                if submission.normalized_data:
                    for field_name, field_value in submission.normalized_data.items():
                        # Sanitize field names for Excel
                        clean_field_name = self._sanitize_column_name(field_name)
                        row[clean_field_name] = field_value
                
                data_rows.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(data_rows)
            
            # Sort by submission date
            if 'Submitted At' in df.columns:
                df = df.sort_values('Submitted At', ascending=False)
            
            # Apply column ordering if specified
            column_order = self.template_config.get('column_order', [])
            if column_order:
                # Reorder columns based on configuration
                available_columns = df.columns.tolist()
                ordered_columns = []
                
                # Add specified columns in order
                for col in column_order:
                    if col in available_columns:
                        ordered_columns.append(col)
                        available_columns.remove(col)
                
                # Add remaining columns
                ordered_columns.extend(available_columns)
                df = df[ordered_columns]
            
            return df
            
        except Exception as e:
            logger.error(f"Error creating DataFrame: {str(e)}")
            raise e
    
    def _create_basic_excel(self, df: pd.DataFrame, file_path: str):
        """Create basic Excel file with pandas"""
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Main data sheet
                df.to_excel(writer, sheet_name='Form Submissions', index=False)
                
                # Summary sheet
                summary_df = self._create_summary_dataframe()
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Data source details sheet
                sources_df = self._create_data_sources_dataframe()
                sources_df.to_excel(writer, sheet_name='Data Sources', index=False)
                
                # Apply basic formatting
                workbook = writer.book
                self._apply_basic_formatting(workbook)
                
        except Exception as e:
            logger.error(f"Error creating basic Excel file: {str(e)}")
            raise e
    
    def _create_advanced_excel(self, df: pd.DataFrame, file_path: str):
        """Create advanced Excel file with custom formatting"""
        try:
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)
            
            # Create main data sheet
            self._create_formatted_data_sheet(wb, df)
            
            # Create summary sheet
            self._create_summary_sheet(wb)
            
            # Create charts sheet if configured
            if self.template_config.get('include_charts'):
                self._create_charts_sheet(wb, df)
            
            # Create data sources sheet
            self._create_data_sources_sheet(wb)
            
            # Save workbook
            wb.save(file_path)
            
        except Exception as e:
            logger.error(f"Error creating advanced Excel file: {str(e)}")
            raise e
    
    def _create_formatted_data_sheet(self, workbook: Workbook, df: pd.DataFrame):
        """Create formatted data sheet with advanced styling"""
        ws = workbook.create_sheet("Form Submissions")
        
        # Add title
        title_cell = ws.cell(row=1, column=1, value=f"Form Submissions Export - {self.export_config.name}")
        title_cell.font = Font(size=16, bold=True, color="2F4F4F")
        
        # Add export info
        info_row = 3
        ws.cell(row=info_row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=info_row + 1, column=1, value=f"Total Submissions: {len(self.submissions)}")
        ws.cell(row=info_row + 2, column=1, value=f"Date Range: {self._get_date_range_text()}")
        
        # Add data starting from row 6
        data_start_row = 6
        
        # Add headers
        for col_idx, column_name in enumerate(df.columns, 1):
            cell = ws.cell(row=data_start_row, column=col_idx, value=column_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data rows
        for row_idx, (_, row) in enumerate(df.iterrows(), data_start_row + 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Format dates
                if isinstance(value, datetime):
                    cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                
                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add table formatting
        if len(df) > 0:
            table_range = f"A{data_start_row}:{ws.cell(row=data_start_row + len(df), column=len(df.columns)).coordinate}"
            table = Table(displayName="FormSubmissionsTable", ref=table_range)
            style = TableStyleInfo(
                name="TableStyleMedium9", 
                showFirstColumn=False,
                showLastColumn=False, 
                showRowStripes=True, 
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)
    
    def _create_summary_sheet(self, workbook: Workbook):
        """Create summary statistics sheet"""
        ws = workbook.create_sheet("Summary")
        
        # Title
        title_cell = ws.cell(row=1, column=1, value="Export Summary")
        title_cell.font = Font(size=14, bold=True)
        
        # Summary statistics
        summary_data = self._calculate_summary_stats()
        
        row = 3
        for key, value in summary_data.items():
            ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _create_charts_sheet(self, workbook: Workbook, df: pd.DataFrame):
        """Create charts and visualizations sheet"""
        ws = workbook.create_sheet("Charts")
        
        # Title
        title_cell = ws.cell(row=1, column=1, value="Data Visualizations")
        title_cell.font = Font(size=14, bold=True)
        
        # Add some basic charts data (simplified for now)
        # In a full implementation, you would use openpyxl's chart features
        
        # Submissions by data source
        if 'Data Source' in df.columns:
            source_counts = df['Data Source'].value_counts()
            
            row = 3
            ws.cell(row=row, column=1, value="Submissions by Data Source").font = Font(bold=True)
            row += 1
            
            for source, count in source_counts.items():
                ws.cell(row=row, column=1, value=source)
                ws.cell(row=row, column=2, value=count)
                row += 1
    
    def _create_data_sources_sheet(self, workbook: Workbook):
        """Create data sources information sheet"""
        ws = workbook.create_sheet("Data Sources")
        
        # Title
        title_cell = ws.cell(row=1, column=1, value="Data Sources Information")
        title_cell.font = Font(size=14, bold=True)
        
        # Headers
        headers = ["Name", "Type", "Source ID", "Last Sync", "Total Submissions"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
        
        # Data source information
        data_source_ids = self.export_config.data_sources
        from ..models import FormDataSource, FormDataSubmission
        
        row = 4
        for ds_id in data_source_ids:
            data_source = FormDataSource.query.get(ds_id)
            if data_source:
                submission_count = FormDataSubmission.query.filter_by(
                    data_source_id=ds_id
                ).count()
                
                ws.cell(row=row, column=1, value=data_source.name)
                ws.cell(row=row, column=2, value=data_source.source_type)
                ws.cell(row=row, column=3, value=data_source.source_id)
                ws.cell(row=row, column=4, value=data_source.last_sync)
                ws.cell(row=row, column=5, value=submission_count)
                row += 1
        
        # Auto-adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 20
    
    def _create_summary_dataframe(self) -> pd.DataFrame:
        """Create summary statistics DataFrame"""
        try:
            summary_data = self._calculate_summary_stats()
            
            summary_df = pd.DataFrame([
                {'Metric': key, 'Value': value} 
                for key, value in summary_data.items()
            ])
            
            return summary_df
            
        except Exception as e:
            logger.error(f"Error creating summary DataFrame: {str(e)}")
            return pd.DataFrame()
    
    def _create_data_sources_dataframe(self) -> pd.DataFrame:
        """Create data sources information DataFrame"""
        try:
            from ..models import FormDataSource, FormDataSubmission
            
            data_sources_data = []
            
            for ds_id in self.export_config.data_sources:
                data_source = FormDataSource.query.get(ds_id)
                if data_source:
                    submission_count = FormDataSubmission.query.filter_by(
                        data_source_id=ds_id
                    ).count()
                    
                    data_sources_data.append({
                        'Name': data_source.name,
                        'Type': data_source.source_type,
                        'Source ID': data_source.source_id,
                        'Last Sync': data_source.last_sync,
                        'Total Submissions': submission_count,
                        'Auto Sync': data_source.auto_sync,
                        'Sync Interval (min)': data_source.sync_interval // 60
                    })
            
            return pd.DataFrame(data_sources_data)
            
        except Exception as e:
            logger.error(f"Error creating data sources DataFrame: {str(e)}")
            return pd.DataFrame()
    
    def _calculate_summary_stats(self) -> Dict[str, Any]:
        """Calculate summary statistics"""
        try:
            total_submissions = len(self.submissions)
            
            # Group by data source
            source_counts = {}
            date_range = {'earliest': None, 'latest': None}
            
            for submission in self.submissions:
                # Count by source
                source_name = submission.data_source.name
                source_counts[source_name] = source_counts.get(source_name, 0) + 1
                
                # Track date range
                submit_date = submission.submitted_at
                if not date_range['earliest'] or submit_date < date_range['earliest']:
                    date_range['earliest'] = submit_date
                if not date_range['latest'] or submit_date > date_range['latest']:
                    date_range['latest'] = submit_date
            
            # Most active source
            most_active_source = max(source_counts.items(), key=lambda x: x[1]) if source_counts else ("None", 0)
            
            summary = {
                'Total Submissions': total_submissions,
                'Data Sources Count': len(set(s.data_source_id for s in self.submissions)),
                'Most Active Source': f"{most_active_source[0]} ({most_active_source[1]} submissions)",
                'Date Range': self._get_date_range_text(),
                'Export Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Export Name': self.export_config.name,
                'Export Description': self.export_config.description or 'No description'
            }
            
            return summary
            
        except Exception as e:
            logger.error(f"Error calculating summary stats: {str(e)}")
            return {'Error': 'Failed to calculate statistics'}
    
    def _apply_basic_formatting(self, workbook: Workbook):
        """Apply basic formatting to Excel sheets"""
        try:
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                
                # Format headers (first row)
                if ws.max_row > 0:
                    for cell in ws[1]:
                        if cell.value:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                    
        except Exception as e:
            logger.error(f"Error applying basic formatting: {str(e)}")
    
    def _sanitize_column_name(self, column_name: str) -> str:
        """Sanitize column names for Excel compatibility"""
        if not isinstance(column_name, str):
            column_name = str(column_name)
        
        # Replace problematic characters
        sanitized = column_name.replace('\n', ' ').replace('\r', ' ')
        
        # Limit length
        if len(sanitized) > 255:  # Excel column name limit
            sanitized = sanitized[:255]
        
        return sanitized
    
    def _get_date_range_text(self) -> str:
        """Get human-readable date range text"""
        try:
            if self.export_config.date_range_start and self.export_config.date_range_end:
                start = self.export_config.date_range_start.strftime('%Y-%m-%d')
                end = self.export_config.date_range_end.strftime('%Y-%m-%d')
                return f"{start} to {end}"
            elif self.export_config.date_range_start:
                start = self.export_config.date_range_start.strftime('%Y-%m-%d')
                return f"From {start}"
            elif self.export_config.date_range_end:
                end = self.export_config.date_range_end.strftime('%Y-%m-%d')
                return f"Until {end}"
            else:
                return "All available data"
        except:
            return "Date range not specified"
